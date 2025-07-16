# Copyright 2021 - 2024 Universität Tübingen, DKFZ, EMBL, and Universität zu Köln
# for the German Human Genome-Phenome Archive (GHGA)
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
#

"""Step definitions for downloading metadata artifacts with the GHGA Data Steward Kit"""

from io import BytesIO
from tempfile import NamedTemporaryFile

from openpyxl import load_workbook

from .conftest import JointFixture, parse, scenarios, then, when

scenarios("../features/140_download_metadata.feature")


@when(
    parse('metadata of dataset "{dataset_alias}" is downloaded from the system'),
    target_fixture="response",
)
def download_metadata(fixtures: JointFixture, dataset_alias: str):
    """Download metadata from the system."""
    dataset_accessions = fixtures.state.get_state("dataset_accessions")
    assert dataset_alias in dataset_accessions, (
        f"Dataset alias '{dataset_alias}' not found in state."
    )
    study_accession = dataset_accessions[dataset_alias]["study_accession"]
    url = f"{fixtures.config.rts_url}/studies/{study_accession}"
    return fixtures.http.get(url)


@then(
    parse(
        'the downloaded spreadsheet should match the expected for dataset "{dataset_alias}"'
    )
)
def check_downloaded_spreadsheet(fixtures: JointFixture, response, dataset_alias: str):
    """Check if the downloaded spreadsheet matches the expected for the dataset."""
    dataset_accessions = fixtures.state.get_state("dataset_accessions")
    assert dataset_alias in dataset_accessions, (
        f"Dataset alias '{dataset_alias}' not found in state."
    )
    study_accession = dataset_accessions[dataset_alias]["study_accession"]

    assert isinstance(response.content, bytes)
    assert response.headers["Content-Disposition"] == (
        f'attachment; filename="{study_accession}.xlsx"'
    )

    with NamedTemporaryFile(delete=False) as temp_file:
        temp_file.write(response.content)
        print(temp_file.name)

    expected_filename = fixtures.config.rts_expected_files[dataset_alias]
    expected_spreadsheet = f"{fixtures.config.data_dir}/metadata/{expected_filename}"

    actual = load_workbook(BytesIO(response.content), read_only=True)
    expected = load_workbook(expected_spreadsheet, read_only=True)

    assert sorted(expected.sheetnames) == sorted(actual.sheetnames), (
        "Sheet names do not match."
    )

    for sheet_name in expected.sheetnames:
        expected_sheet = expected[sheet_name]
        actual_sheet = actual[sheet_name]

        expected_data = [
            [
                str(cell.value) if not str(cell.value).startswith("GHGA") else "GHGA"
                for cell in row
            ]
            for row in expected_sheet.iter_rows()
        ]
        actual_data = [
            [
                str(cell.value) if not str(cell.value).startswith("GHGA") else "GHGA"
                for cell in row
            ]
            for row in actual_sheet.iter_rows()
        ]

        # Sort both datasets for comparison
        expected_data_sorted = sorted([sorted(d) for d in expected_data])
        actual_data_sorted = sorted([sorted(d) for d in actual_data])

        assert expected_data_sorted == actual_data_sorted, (
            f"Data mismatch in sheet: {sheet_name}. Expected: {expected_data_sorted},"
            + f" Got: {actual_data_sorted}"
        )

    return
